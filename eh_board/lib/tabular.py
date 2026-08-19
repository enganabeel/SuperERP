# -*- encoding: utf-8 -*-
##############################################################################
#
# ERP Heritage
# Copyright (C) 2026 (https://www.erpheritage.com.au/)
#
##############################################################################
"""Pure-Python tabular parsing and in-memory aggregation for file data sources.

The design is the deliberate opposite of the incumbent's: an uploaded file is
NEVER turned into an Odoo model or a physical Postgres table. It is parsed to a
normalised list of dict-rows plus a small column manifest, cached as one JSON
attachment, and aggregated in Python (the datasets are capped and small). No
``ir.model``, no ``ir.model.fields``, no schema pollution, no global scratch
tables, no per-column manual typing.

This module holds no Odoo model and imports nothing from Odoo, so it can be unit
tested without a database. CSV needs zero third-party dependencies; the Excel
path imports ``openpyxl`` lazily and fails with a friendly message if absent.
"""
import csv
import io
import math
from datetime import date, datetime

# Hard caps so a large upload can never OOM a worker or stall a render.
MAX_ROWS = 50000
MAX_COLS = 60

# Column data types we distinguish. Kept intentionally small.
DTYPES = ("number", "date", "bool", "text")

_TRUE = {"true", "yes", "y", "1", "t"}
_FALSE = {"false", "no", "n", "0", "f"}

_DATE_FORMATS = (
    "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y",
    "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%d/%m/%Y %H:%M",
    "%d %b %Y", "%d %B %Y", "%b %d, %Y", "%B %d, %Y",
)


class TabularError(Exception):
    """Raised when a file cannot be read as a table."""


# --------------------------------------------------------------------------
# Cell coercion + type inference
# --------------------------------------------------------------------------
def _try_number(text):
    """Return a float for a numeric cell, else None. Tolerates thousands
    separators and a leading currency symbol / trailing %."""
    s = (text or "").strip()
    if not s:
        return None
    neg = False
    if s.startswith("(") and s.endswith(")"):  # accounting negatives
        neg, s = True, s[1:-1]
    s = s.lstrip("$€£¥ ").rstrip("% ").replace(",", "").strip()
    if not s:
        return None
    # Preserve identifier-like codes (zip, SKU, store id, phone): a leading-zero
    # all-digit string is a code, not a quantity. Coercing "007" to 7.0 drops the
    # zero and collapses distinct codes into one group.
    if len(s) > 1 and s[0] == "0" and s.isdigit():
        return None
    try:
        val = float(s)
    except ValueError:
        return None
    # Reject inf / nan / overflow (e.g. '1e400' -> inf). A non-finite value that
    # reached an aggregate would serialise as a bare NaN/Infinity token and break
    # the whole board's JSON-RPC response, not just one widget.
    if not math.isfinite(val):
        return None
    return -val if neg else val


def _try_date(text, fmt=None):
    """Return a date/datetime for a recognised date cell, else None.

    When ``fmt`` is given (the column's detected format), it is tried first so a
    whole column parses consistently - the fix for the day/month vs month/day
    ambiguity that silently swapped US-formatted dates."""
    s = (text or "").strip()
    if not s or len(s) > 32:
        return None
    formats = ([fmt] + list(_DATE_FORMATS)) if fmt else _DATE_FORMATS
    for f in formats:
        try:
            return datetime.strptime(s, f)
        except ValueError:
            continue
    return None


def _infer_date_format(samples):
    """Pick one date format for a whole column so it parses consistently.

    Returns the first format that parses EVERY sampled cell; a column with any
    disambiguating value (a day > 12) therefore resolves day/month vs month/day
    correctly for the whole column instead of guessing per cell. Falls back to
    the format that parses the most cells."""
    seen = [(s or "").strip() for s in samples if (s or "").strip()][:200]
    if not seen:
        return None
    best, best_n = None, -1
    for fmt in _DATE_FORMATS:
        n = 0
        for s in seen:
            try:
                datetime.strptime(s, fmt)
                n += 1
            except ValueError:
                pass
        if n == len(seen):
            return fmt
        if n > best_n:
            best, best_n = fmt, n
    return best


def _try_bool(text):
    s = (text or "").strip().lower()
    if s in _TRUE:
        return True
    if s in _FALSE:
        return False
    return None


def _infer_dtype(samples):
    """Infer a column dtype from a sample of raw string cells."""
    seen = [s for s in samples if (s or "").strip()]
    if not seen:
        return "text"
    # Number BEFORE bool: a 0/1 flag column is numeric (so it sums/averages),
    # not boolean. Genuine true/false/yes/no columns fail number inference and
    # fall through to the bool test.
    if all(_try_number(s) is not None for s in seen):
        return "number"
    if all(_try_bool(s) is not None for s in seen):
        return "bool"
    if all(_try_date(s) is not None for s in seen):
        return "date"
    return "text"


def _coerce(value, dtype, date_format=None):
    """Coerce a raw string cell to the column's dtype (None on blank/mismatch)."""
    if value is None:
        return None
    if dtype == "number":
        return _try_number(value)
    if dtype == "bool":
        return _try_bool(value)
    if dtype == "date":
        d = _try_date(value, date_format)
        return d.isoformat() if d else None
    s = str(value).strip()
    return s or None


def _slug(header, taken):
    """Stable, unique column name from a header cell."""
    base = "".join(ch.lower() if ch.isalnum() else "_" for ch in (header or "").strip())
    base = "_".join(filter(None, base.split("_"))) or "col"
    name, i = base, 1
    while name in taken:
        i += 1
        name = "%s_%d" % (base, i)
    taken.add(name)
    return name


# --------------------------------------------------------------------------
# Parsers
# --------------------------------------------------------------------------
def _finish(header_cells, raw_rows):
    """Build the column manifest + coerced rows from a header + raw string rows."""
    if not header_cells:
        raise TabularError("The file has no header row.")
    header_cells = header_cells[:MAX_COLS]
    taken = set()
    columns = []
    for cell in header_cells:
        label = (str(cell).strip() if cell is not None else "") or ("Column %d" % (len(columns) + 1))
        columns.append({"name": _slug(label, taken), "label": label})

    truncated = len(raw_rows) > MAX_ROWS
    raw_rows = raw_rows[:MAX_ROWS]

    ncol = len(columns)
    # Sample the first ~200 rows per column to infer dtype (and, for a date
    # column, one consistent format for the whole column).
    for ci, col in enumerate(columns):
        samples = [(r[ci] if ci < len(r) else "") for r in raw_rows[:200]]
        col["dtype"] = _infer_dtype(samples)
        col["date_format"] = _infer_date_format(samples) if col["dtype"] == "date" else None

    rows = []
    for r in raw_rows:
        row = {}
        empty = True
        for ci, col in enumerate(columns):
            raw = r[ci] if ci < len(r) else ""
            val = _coerce(raw, col["dtype"], col.get("date_format"))
            if val is not None:
                empty = False
            row[col["name"]] = val
        if not empty:
            rows.append(row)
    return {"columns": columns, "rows": rows, "row_count": len(rows),
            "truncated": truncated}


def parse_csv(raw_bytes):
    """Parse CSV/TSV bytes into a manifest + coerced rows (stdlib only)."""
    if not raw_bytes:
        raise TabularError("The file is empty.")
    text = None
    # A byte-order mark pins the encoding: without this, a UTF-16/32 file would
    # be silently mis-decoded by latin-1 (which never raises) into garbage
    # columns. Use the GENERIC utf-16 / utf-32 codecs (not -le/-be), which consume
    # the BOM instead of leaving a stray U+FEFF on the first column header.
    encodings = ["utf-8-sig", "utf-8", "latin-1"]
    for bom, enc in ((b"\xff\xfe\x00\x00", "utf-32"), (b"\x00\x00\xfe\xff", "utf-32"),
                     (b"\xff\xfe", "utf-16"), (b"\xfe\xff", "utf-16")):
        if raw_bytes.startswith(bom):
            encodings = [enc] + encodings
            break
    for enc in encodings:
        try:
            text = raw_bytes.decode(enc)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    if text is None:
        raise TabularError("The file's text encoding could not be read.")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    # Bounded read: stop a few rows past the cap so a huge upload never loads
    # wholesale into memory (parse_xlsx already caps during iteration).
    all_rows = []
    try:
        for row in reader:
            all_rows.append(row)
            if len(all_rows) > MAX_ROWS + 1:
                break
    except csv.Error as err:
        raise TabularError("This file could not be read as CSV (%s)." % err)
    if not all_rows:
        raise TabularError("The file has no rows.")
    return _finish(all_rows[0], all_rows[1:])


def parse_xlsx(raw_bytes):
    """Parse the first sheet of an .xlsx workbook (lazy openpyxl)."""
    try:
        import openpyxl  # noqa: F401 - optional dependency
    except ImportError:
        raise TabularError(
            "Reading Excel needs the 'openpyxl' library on the server; save the "
            "file as CSV, or ask your administrator to install openpyxl.")
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception as err:  # noqa: BLE001 - normalise to a friendly error
        raise TabularError("This file could not be read as an Excel workbook.")
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if c is None else str(c) for c in row])
        if len(rows) > MAX_ROWS + 1:
            break
    wb.close()
    if not rows:
        raise TabularError("The workbook's first sheet is empty.")
    return _finish(rows[0], rows[1:])


def parse(raw_bytes, kind):
    """Dispatch to the CSV or Excel parser by ``kind`` (csv|xlsx)."""
    if kind == "xlsx":
        return parse_xlsx(raw_bytes)
    return parse_csv(raw_bytes)


# --------------------------------------------------------------------------
# In-memory aggregation (same normalised shape as aggregation.aggregate)
# --------------------------------------------------------------------------
def _bucket_date(iso, granularity):
    """Bucket an ISO date/datetime string to a coarser period label."""
    if not iso:
        return None
    try:
        dt = datetime.fromisoformat(iso)
    except ValueError:
        return iso
    g = granularity or "month"
    if g == "year":
        return "%04d" % dt.year
    if g == "quarter":
        return "%04d-Q%d" % (dt.year, (dt.month - 1) // 3 + 1)
    if g == "week":
        iso_cal = dt.isocalendar()
        return "%04d-W%02d" % (iso_cal[0], iso_cal[1])
    if g == "day":
        return dt.strftime("%Y-%m-%d")
    if g == "hour":
        return dt.strftime("%Y-%m-%d %H:00")
    if g == "minute":
        return dt.strftime("%Y-%m-%d %H:%M")
    if g == "month_number":
        return "%02d" % dt.month
    if g == "day_of_week":
        return dt.strftime("%A")
    # default: month
    return dt.strftime("%Y-%m")


def _dim_value(row, dim):
    val = row.get(dim["field"])
    if dim.get("dtype") == "date":
        return _bucket_date(val, dim.get("granularity"))
    return val


def aggregate_records(rows, dimensions, measures):
    """Group ``rows`` (list of dicts) by the dimension columns and reduce each
    measure. Returns the same normalised structure as aggregation.aggregate."""
    measure_meta = []
    for m in measures:
        measure_meta.append({
            "key": m["key"], "field": m.get("field"),
            "verb": m.get("verb", "count"),
            "multiplier": m.get("multiplier", 1.0) or 1.0})

    groups = {}     # gkey -> {"keys","labels","acc":{mkey: aggregator-state}}
    order = []
    for row in rows:
        dim_vals = [_dim_value(row, d) for d in dimensions]
        labels = ["" if v is None else str(v) for v in dim_vals]
        gkey = tuple(labels)
        g = groups.get(gkey)
        if g is None:
            g = {"keys": list(dim_vals), "labels": labels, "acc": {}}
            groups[gkey] = g
            order.append(gkey)
        for meta in measure_meta:
            _accumulate(g["acc"], meta, row)

    out_rows = []
    for gkey in order:
        g = groups[gkey]
        values = {}
        for meta in measure_meta:
            values[meta["key"]] = _finalize(g["acc"].get(meta["key"]), meta) \
                * meta["multiplier"]
        out_rows.append({"keys": g["keys"], "labels": g["labels"], "values": values})

    return {
        "rows": out_rows,
        "measures": [m["key"] for m in measures],
        "dimensions": [d["field"] for d in dimensions],
        "measure_verbs": {m["key"]: m.get("verb", "count") for m in measures},
    }


def _accumulate(acc, meta, row):
    key, verb, field = meta["key"], meta["verb"], meta["field"]
    state = acc.get(key)
    if verb == "count":
        acc[key] = (state or 0) + 1
        return
    raw = row.get(field) if field else None
    if verb == "count_distinct":
        s = state or set()
        if raw is not None:
            s.add(raw if not isinstance(raw, (list, dict)) else str(raw))
        acc[key] = s
        return
    num = raw if isinstance(raw, (int, float)) and not isinstance(raw, bool) else None
    if num is None:
        # Non-numeric cell under a numeric verb: count non-empty as a fallback.
        if verb == "sum":
            acc[key] = (state or 0.0)
        else:
            acc[key] = state
        return
    if verb == "sum":
        acc[key] = (state or 0.0) + num
    elif verb == "avg":
        s = state or {"sum": 0.0, "n": 0}
        s["sum"] += num
        s["n"] += 1
        acc[key] = s
    elif verb == "min":
        acc[key] = num if state is None else min(state, num)
    elif verb == "max":
        acc[key] = num if state is None else max(state, num)
    else:
        acc[key] = (state or 0.0) + num


def _finalize(state, meta):
    verb = meta["verb"]
    if state is None:
        return 0.0
    if verb == "count":
        return float(state)
    if verb == "count_distinct":
        return float(len(state))
    if verb == "avg":
        return (state["sum"] / state["n"]) if state.get("n") else 0.0
    return float(state)
